from flask import Flask, request, render_template_string, flash, redirect, url_for
import openpyxl

app = Flask(__name__)
app.secret_key = "1234"  # Required for flashing messages

# Path to the Excel file
EXCEL_FILE = "data.xlsx"
START_ROW = 2  # Start from row 2
LABEL_COLUMN = "B"  # Column for labels
DATA_COLUMN = "M"  # Column for data

def write_to_excel(label, value):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Search for the label in the label column
        row = START_ROW
        while ws[f"{LABEL_COLUMN}{row}"].value:
            if ws[f"{LABEL_COLUMN}{row}"].value == label:
                # Overwrite the existing data in column M (same row)
                ws[f"{DATA_COLUMN}{row}"].value = value
                
                wb.save(EXCEL_FILE)
                wb.close()
                return "Success"
            row += 1
        
        # If label not found, add a new row
        ws[f"{LABEL_COLUMN}{row}"] = label
        ws[f"{DATA_COLUMN}{row}"] = value
        
        wb.save(EXCEL_FILE)
        wb.close()
    except Exception as e:
        return str(e)
    return "Success"

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        label = request.form.get("label")
        data = request.form.get("data")
        if label and data:
            message = write_to_excel(label, data)
            flash(message, "success" if message == "Success" else "error")
            return redirect(url_for('index'))

    return render_template_string('''
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Data Entry Form</title>
            <style>
                body {
                    display: flex;
                    justify-content: center;
                    align-items: center;
                    height: 100vh;
                    background-color: #516B78;
                    font-family: Arial, sans-serif;
                }
                .container {
                    background: white;
                    padding: 20px;
                    border-radius: 10px;
                    box-shadow: 0 0 10px rgba(0, 0, 0, 0.1);
                    text-align: center;
                }
                .input-group {
                    display: flex;
                    flex-direction: column;
                    margin-bottom: 15px;
                }
                label {
                    font-weight: bold;
                    margin-bottom: 5px;
                }
                input {
                    padding: 10px;
                    border: 1px solid #313638;
                    border-radius: 5px;
                    width: 300px;
                }
                .form-row {
                    display: flex;
                    align-items: center;
                    gap: 10px;
                }
                button {
                    padding: 10px 20px;
                    border: none;
                    background-color: #00D043;
                    color: white;
                    font-size: 16px;
                    border-radius: 5px;
                    cursor: pointer;
                }
                button:hover {
                    background-color: #218838;
                }
                .message {
                    font-size: 16px;
                    margin-top: 10px;
                }
                .success {
                    color: green;
                }
                .error {
                    color: red;
                }
            </style>
        </head>
        <body>
            <div class="container">
                <h2>Data Entry Form</h2>
                <form method="post">
                    <div class="input-group">
                        <label for="label">Label</label>
                        <input type="text" id="label" name="label" required>
                    </div>
                    <div class="form-row">
                        <input type="text" name="data" placeholder="Enter Data" required>
                        <button type="submit">Submit</button>
                    </div>
                </form>
                {% with messages = get_flashed_messages(with_categories=true) %}
                    {% if messages %}
                        <div class="message">
                            {% for category, message in messages %}
                                <p class="{{ category }}">{{ message }}</p>
                            {% endfor %}
                        </div>
                    {% endif %}
                {% endwith %}
            </div>
        </body>
        </html>
    ''')

if __name__ == "__main__":
    app.run(debug=True)
